"""人员工作簿导入的纯校验、匹配、加密写入与凭据文件安全测试。"""
from __future__ import annotations

import os
import stat
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from app.services import personnel_import_service as service
from scripts.build_personnel_template import build_template


DICT_DIR = Path(__file__).resolve().parents[1] / "data" / "tj_dicts"


def _id_card(prefix17: str) -> str:
    total = sum(int(prefix17[i]) * service._ID_WEIGHTS[i] for i in range(17))
    return prefix17 + service._ID_CHECKS[total % 11]


DOCTOR_ID = _id_card("11010519900101001")
PHARMACIST_ID = _id_card("11010519920202002")


def _set_row(ws, row_number: int, values: dict[str, object]) -> None:
    headers = {
        str(cell.value).replace(" *", ""): cell.column
        for cell in ws[1]
    }
    for name, value in values.items():
        ws.cell(row_number, headers[name], value)


def _filled_workbook(tmp_path: Path) -> Path:
    path = build_template(tmp_path / "人员.xlsx")
    wb = load_workbook(path)
    _set_row(wb["医生"], 2, {
        "手机号": "13900000001",
        "姓名": "测试医师",
        "身份证号": DOCTOR_ID,
        "医师资格证号": "TEST-L-001",
        "医师执业证号": "TEST-P-001",
        "科室名称": "内科",
        "科室类别编码": "03",
        "诊疗科目编码": "03",
        "诊疗科目名称": "内科",
        "职称": "主治医师",
        "执业年限": "10",
        "擅长": "测试专长",
        "挂号费（元）": "0.01",
        "医务审核结论": "通过",
        "CA名单核对": "是",
    })
    _set_row(wb["药师"], 2, {
        "登录用户名": "pharm_test",
        "姓名": "测试药师",
        "身份证号": PHARMACIST_ID,
        "手机号": "13800000001",
        "启用状态": "启用",
        "资质审核结论": "通过",
        "CA名单核对": "是",
    })
    wb.save(path)
    wb.close()
    return path


def test_id_card_checksum_and_date():
    assert service.valid_id_card(DOCTOR_ID)
    assert not service.valid_id_card(DOCTOR_ID[:-1] + ("1" if DOCTOR_ID[-1] != "1" else "2"))
    assert not service.valid_id_card("110105199013010010")


def test_rejects_invalid_xlsx_container(tmp_path):
    path = tmp_path / "invalid.xlsx"
    path.write_bytes(b"not-an-xlsx")
    data = service.parse_personnel_workbook(path, DICT_DIR)
    assert data.errors
    assert "有效的 xlsx" in data.errors[0].message


def test_parse_valid_workbook(tmp_path):
    data = service.parse_personnel_workbook(_filled_workbook(tmp_path), DICT_DIR)
    assert not data.errors
    assert len(data.doctors) == 1
    assert len(data.pharmacists) == 1
    assert data.doctors[0].register_fee_yuan == service.Decimal("0.01")
    assert data.pharmacists[0].active is True


def test_parse_rejects_cross_sheet_duplicate_without_echoing_value(tmp_path):
    path = _filled_workbook(tmp_path)
    wb = load_workbook(path)
    _set_row(wb["药师"], 2, {"身份证号": DOCTOR_ID})
    wb.save(path)
    wb.close()

    data = service.parse_personnel_workbook(path, DICT_DIR)
    rendered = "\n".join(issue.render() for issue in data.issues)
    assert data.errors
    assert "重复" in rendered
    assert DOCTOR_ID not in rendered
    assert "测试医师" not in rendered
    assert "测试药师" not in rendered


def test_parse_rejects_numeric_sensitive_cell_without_echoing_value(tmp_path):
    path = _filled_workbook(tmp_path)
    wb = load_workbook(path)
    _set_row(wb["医生"], 2, {"身份证号": 110105199001010010})
    wb.save(path)
    wb.close()

    data = service.parse_personnel_workbook(path, DICT_DIR)
    rendered = "\n".join(issue.render() for issue in data.issues)
    assert "必须设置为文本格式" in rendered
    assert "110105199001010010" not in rendered


def test_parse_rejects_formulas(tmp_path):
    path = _filled_workbook(tmp_path)
    wb = load_workbook(path)
    _set_row(wb["医生"], 2, {"姓名": '=HYPERLINK("https://invalid.example","x")'})
    wb.save(path)
    wb.close()

    data = service.parse_personnel_workbook(path, DICT_DIR)
    rendered = "\n".join(issue.render() for issue in data.issues)
    assert data.errors
    assert "不允许使用公式" in rendered
    assert "invalid.example" not in rendered


def test_database_plan_matches_doctor_phone_and_creates_pharmacist(tmp_path, monkeypatch):
    data = service.parse_personnel_workbook(_filled_workbook(tmp_path), DICT_DIR)
    monkeypatch.setattr(service, "decrypt", lambda value: value)
    doctor = SimpleNamespace(id=7, id_card_enc=None)
    user = SimpleNamespace(id=70, phone_enc="13900000001")

    plan = service.build_database_plan(data, [(doctor, user)], [])
    assert not plan.errors
    assert len(plan.doctors) == 1
    assert plan.doctors[0].target is doctor
    assert plan.new_pharmacists == 1


def test_database_plan_refuses_fake_doctor_account(tmp_path, monkeypatch):
    data = service.parse_personnel_workbook(_filled_workbook(tmp_path), DICT_DIR)
    monkeypatch.setattr(service, "decrypt", lambda value: value)
    plan = service.build_database_plan(data, [], [])
    assert plan.errors
    rendered = "\n".join(issue.render() for issue in plan.issues)
    assert "本人先完成微信手机号授权登录" in rendered
    assert "13900000001" not in rendered


@pytest.mark.asyncio
async def test_apply_plan_encrypts_ids_and_writes_summary_audit(tmp_path, monkeypatch):
    data = service.parse_personnel_workbook(_filled_workbook(tmp_path), DICT_DIR)
    monkeypatch.setattr(service, "decrypt", lambda value: value)
    monkeypatch.setattr(service, "encrypt", lambda value: f"encrypted:{value}")
    monkeypatch.setattr(service, "hash_password", lambda value: f"hashed:{len(value)}")
    doctor = SimpleNamespace(
        id=7, id_card_enc=None, audit_status="pending", years=None, good_at=None,
        register_fee_fen=5000,
    )
    user = SimpleNamespace(id=70, phone_enc="13900000001")
    plan = service.build_database_plan(data, [(doctor, user)], [])

    class FakeDb:
        def __init__(self):
            self.added = []

        def add(self, value):
            self.added.append(value)

        async def flush(self):
            return None

    db = FakeDb()
    result = await service.apply_database_plan(
        db, plan, approve_doctors=True, source_digest="a" * 64
    )
    assert result.doctors_updated == 1
    assert result.doctors_approved == 1
    assert result.pharmacists_created == 1
    assert doctor.audit_status == "approved"
    assert doctor.id_card_enc.startswith("encrypted:")
    assert DOCTOR_ID not in repr(db.added)
    assert any(item.__class__.__name__ == "AuditLog" for item in db.added)
    audit = next(item for item in db.added if item.__class__.__name__ == "AuditLog")
    assert "a" * 64 in audit.detail
    assert len(result.credentials) == 1
    assert PHARMACIST_ID not in repr(result.credentials)


def test_database_plan_blocks_undecryptable_existing_data(tmp_path, monkeypatch):
    data = service.parse_personnel_workbook(_filled_workbook(tmp_path), DICT_DIR)

    def fake_decrypt(value):
        if value == "broken":
            raise ValueError("wrong key")
        return value

    monkeypatch.setattr(service, "decrypt", fake_decrypt)
    doctor = SimpleNamespace(id=7, id_card_enc=None)
    user = SimpleNamespace(id=70, phone_enc="broken")
    plan = service.build_database_plan(data, [(doctor, user)], [])
    rendered = "\n".join(issue.render() for issue in plan.issues)
    assert plan.errors
    assert "已有密文无法用当前密钥解密" in rendered
    assert "broken" not in rendered


def test_temporary_password_and_exclusive_credentials_file(tmp_path):
    password = service.generate_temporary_password()
    assert len(password) == 20
    assert any(c.islower() for c in password)
    assert any(c.isupper() for c in password)
    assert any(c.isdigit() for c in password)
    assert any(c in "!@#$%*-_" for c in password)

    if os.name == "posix":
        os.chmod(tmp_path, 0o700)
    output = tmp_path / "credentials.csv"
    service.write_credentials_file(output, (("pharm_test", password),))
    assert output.exists()
    assert password in output.read_text(encoding="utf-8-sig")
    if os.name == "posix":
        assert stat.S_IMODE(output.stat().st_mode) == 0o600
    with pytest.raises(FileExistsError):
        service.write_credentials_file(output, (("pharm_test", password),))
