"""医师、药师人员工作簿校验与幂等导入业务。

安全边界：
- 错误只包含工作表、行号、字段名和固定原因，不回显单元格内容；
- 医师只能按本人首次微信登录后留存的手机号匹配，禁止伪造 openid；
- 身份证只在进程内短暂存在，写库前使用现有 Fernet 密钥加密；
- 新药师初始密码随机生成，只允许写入独立的 0600 凭据文件。
"""
from __future__ import annotations

import csv
import os
import re
import secrets
import string
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook

from ..core.crypto import decrypt, encrypt
from ..core.security import hash_password
from ..models.audit_log import AuditLog
from ..models.staff import Staff


DOCTOR_SHEET = "医生"
PHARMACIST_SHEET = "药师"
NURSE_SHEET = "护士_暂不导入"
DOCTOR_HEADERS = [
    "手机号", "姓名", "身份证号", "医师资格证号", "医师执业证号", "科室名称",
    "科室类别编码", "诊疗科目编码", "诊疗科目名称", "职称", "执业年限", "擅长",
    "挂号费（元）", "医务审核结论", "CA名单核对", "备注",
]
PHARMACIST_HEADERS = [
    "登录用户名", "姓名", "身份证号", "手机号", "药学资格证号", "执业药师注册证号",
    "科室/部门", "职称", "启用状态", "资质审核结论", "CA名单核对", "备注",
]
DOCTOR_REQUIRED = {
    "手机号", "姓名", "身份证号", "医师资格证号", "医师执业证号", "科室名称",
    "科室类别编码", "诊疗科目编码", "诊疗科目名称", "职称", "医务审核结论", "CA名单核对",
}
PHARMACIST_REQUIRED = {
    "登录用户名", "姓名", "身份证号", "启用状态", "资质审核结论", "CA名单核对",
}
TEXT_ONLY_FIELDS = {
    "手机号", "身份证号", "医师资格证号", "医师执业证号", "科室类别编码",
    "诊疗科目编码", "登录用户名", "药学资格证号", "执业药师注册证号",
}
_PHONE_RE = re.compile(r"^1[3-9]\d{9}$")
_USERNAME_RE = re.compile(r"^[A-Za-z][A-Za-z0-9_.-]{2,63}$")
_ID_RE = re.compile(r"^\d{17}[\dX]$")
_ID_WEIGHTS = (7, 9, 10, 5, 8, 4, 2, 1, 6, 3, 7, 9, 10, 5, 8, 4, 2)
_ID_CHECKS = "10X98765432"
_ALLOWED_SUBJECT_PREFIXES = ("03", "05", "50")
MAX_WORKBOOK_BYTES = 10 * 1024 * 1024
MAX_WORKBOOK_UNCOMPRESSED_BYTES = 50 * 1024 * 1024


@dataclass(frozen=True)
class ImportIssue:
    level: str
    sheet: str
    row: int | None
    field: str | None
    message: str

    def render(self) -> str:
        location = self.sheet
        if self.row is not None:
            location += f" 第{self.row}行"
        if self.field:
            location += f" {self.field}"
        return f"[{self.level}] {location}: {self.message}"


@dataclass(frozen=True)
class DoctorImportRow:
    row_number: int
    phone: str
    name: str
    id_card: str
    license_no: str
    practice_no: str
    dept: str
    dept_code: str
    subject_code: str
    subject_name: str
    title: str
    years: int | None
    good_at: str | None
    register_fee_yuan: Decimal | None
    review_result: str
    ca_matched: bool


@dataclass(frozen=True)
class PharmacistImportRow:
    row_number: int
    username: str
    name: str
    id_card: str
    mobile: str | None
    active: bool
    review_result: str
    ca_matched: bool


@dataclass
class ParsedPersonnel:
    doctors: list[DoctorImportRow] = field(default_factory=list)
    pharmacists: list[PharmacistImportRow] = field(default_factory=list)
    nurse_rows: int = 0
    issues: list[ImportIssue] = field(default_factory=list)

    @property
    def errors(self) -> list[ImportIssue]:
        return [item for item in self.issues if item.level == "ERROR"]


@dataclass(frozen=True)
class DoctorOperation:
    row: DoctorImportRow
    target: Any


@dataclass(frozen=True)
class PharmacistOperation:
    row: PharmacistImportRow
    target: Any | None

    @property
    def is_new(self) -> bool:
        return self.target is None


@dataclass
class DatabasePlan:
    doctors: list[DoctorOperation] = field(default_factory=list)
    pharmacists: list[PharmacistOperation] = field(default_factory=list)
    issues: list[ImportIssue] = field(default_factory=list)

    @property
    def errors(self) -> list[ImportIssue]:
        return [item for item in self.issues if item.level == "ERROR"]

    @property
    def new_pharmacists(self) -> int:
        return sum(op.is_new for op in self.pharmacists)


@dataclass(frozen=True)
class ApplyResult:
    doctors_updated: int
    doctors_approved: int
    pharmacists_created: int
    pharmacists_updated: int
    credentials: tuple[tuple[str, str], ...]


def valid_id_card(value: str) -> bool:
    """校验大陆 18 位身份证格式、出生日期和校验位。"""
    value = (value or "").strip().upper()
    if not _ID_RE.fullmatch(value):
        return False
    try:
        from datetime import datetime

        datetime.strptime(value[6:14], "%Y%m%d")
    except ValueError:
        return False
    total = sum(int(value[i]) * _ID_WEIGHTS[i] for i in range(17))
    return value[-1] == _ID_CHECKS[total % 11]


def generate_temporary_password(length: int = 20) -> str:
    """生成包含大小写、数字和符号的随机初始密码。"""
    if length < 16:
        raise ValueError("临时密码长度不得少于 16")
    alphabet = string.ascii_letters + string.digits + "!@#$%*-_"
    while True:
        value = "".join(secrets.choice(alphabet) for _ in range(length))
        if (
            any(c.islower() for c in value)
            and any(c.isupper() for c in value)
            and any(c.isdigit() for c in value)
            and any(c in "!@#$%*-_" for c in value)
        ):
            return value


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().removesuffix(" *").strip()


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip()
    return str(value).strip()


def _issue(
    issues: list[ImportIssue],
    sheet: str,
    row: int | None,
    field_name: str | None,
    message: str,
    *,
    level: str = "ERROR",
) -> None:
    issues.append(ImportIssue(level, sheet, row, field_name, message))


def _sheet_rows(ws, headers: list[str], required: set[str], issues: list[ImportIssue]):
    header_map: dict[str, int] = {}
    for cell in ws[1]:
        name = _normalize_header(cell.value)
        if name:
            if name in header_map:
                _issue(issues, ws.title, 1, name, "表头重复")
            header_map[name] = cell.column
    for name in headers:
        if name not in header_map:
            _issue(issues, ws.title, 1, name, "缺少模板字段，请使用最新版空白模板")
    if any(item.level == "ERROR" and item.sheet == ws.title and item.row == 1 for item in issues):
        return

    for row_number in range(2, ws.max_row + 1):
        cells = {name: ws.cell(row_number, header_map[name]) for name in headers}
        if not any(cell.value not in (None, "") for cell in cells.values()):
            continue
        values = {name: _text(cell.value) for name, cell in cells.items()}
        for name, cell in cells.items():
            if cell.data_type == "f":
                _issue(issues, ws.title, row_number, name, "不允许使用公式，请粘贴为纯文本值")
        for name in required:
            if not values[name]:
                _issue(issues, ws.title, row_number, name, "必填项为空")
        for name in TEXT_ONLY_FIELDS.intersection(headers):
            cell = cells[name]
            if cell.value not in (None, "") and isinstance(cell.value, (int, float)):
                _issue(issues, ws.title, row_number, name, "必须设置为文本格式，避免前导零或长数字精度丢失")
        yield row_number, values


def _load_code_map(path: Path) -> dict[str, str]:
    with path.open("r", encoding="utf-8-sig", newline="") as fp:
        rows = csv.reader(fp)
        next(rows, None)
        return {str(row[0]).strip(): str(row[1]).strip() for row in rows if len(row) >= 2 and row[0]}


def _optional_int(
    value: str, issues: list[ImportIssue], sheet: str, row: int, field_name: str,
    minimum: int, maximum: int,
) -> int | None:
    if not value:
        return None
    try:
        result = int(value)
    except ValueError:
        _issue(issues, sheet, row, field_name, "必须填写整数")
        return None
    if not minimum <= result <= maximum:
        _issue(issues, sheet, row, field_name, f"必须在 {minimum} 到 {maximum} 之间")
        return None
    return result


def _optional_money(
    value: str, issues: list[ImportIssue], sheet: str, row: int, field_name: str,
) -> Decimal | None:
    if not value:
        return None
    try:
        result = Decimal(value)
    except InvalidOperation:
        _issue(issues, sheet, row, field_name, "必须填写有效金额")
        return None
    if result < 0 or result > Decimal("100000"):
        _issue(issues, sheet, row, field_name, "必须为 0 到 100000 元之间的金额")
        return None
    if result.as_tuple().exponent < -2:
        _issue(issues, sheet, row, field_name, "最多保留两位小数")
        return None
    return result


def _parse_doctors(ws, data: ParsedPersonnel, dept_codes: dict[str, str], subject_codes: dict[str, str]) -> None:
    seen_phones: dict[str, int] = {}
    seen_ids: dict[str, int] = {}
    for row_number, values in _sheet_rows(ws, DOCTOR_HEADERS, DOCTOR_REQUIRED, data.issues) or []:
        phone = values["手机号"]
        id_card = values["身份证号"].upper()
        if phone and not _PHONE_RE.fullmatch(phone):
            _issue(data.issues, DOCTOR_SHEET, row_number, "手机号", "格式不正确")
        if id_card and not valid_id_card(id_card):
            _issue(data.issues, DOCTOR_SHEET, row_number, "身份证号", "格式、出生日期或校验位不正确")
        if phone in seen_phones:
            _issue(data.issues, DOCTOR_SHEET, row_number, "手机号", f"与第{seen_phones[phone]}行重复")
        elif phone:
            seen_phones[phone] = row_number
        if id_card in seen_ids:
            _issue(data.issues, DOCTOR_SHEET, row_number, "身份证号", f"与第{seen_ids[id_card]}行重复")
        elif id_card:
            seen_ids[id_card] = row_number

        dept_code = values["科室类别编码"]
        subject_code = values["诊疗科目编码"]
        subject_name = values["诊疗科目名称"]
        if dept_code and dept_code not in dept_codes:
            _issue(data.issues, DOCTOR_SHEET, row_number, "科室类别编码", "不在天津监管字典 3.2 中")
        if subject_code and subject_code not in subject_codes:
            _issue(data.issues, DOCTOR_SHEET, row_number, "诊疗科目编码", "不在天津监管字典 3.1 中")
        elif subject_code and not subject_code.startswith(_ALLOWED_SUBJECT_PREFIXES):
            _issue(data.issues, DOCTOR_SHEET, row_number, "诊疗科目编码", "超出本院当前许可范围 03/05/50")
        if subject_code in subject_codes and subject_name and subject_codes[subject_code] != subject_name:
            _issue(data.issues, DOCTOR_SHEET, row_number, "诊疗科目名称", "与所选编码的官方字典名称不一致")

        review = values["医务审核结论"]
        ca_matched = values["CA名单核对"] == "是"
        if review not in {"待审核", "通过", "不通过"}:
            _issue(data.issues, DOCTOR_SHEET, row_number, "医务审核结论", "必须从模板下拉值中选择")
        if values["CA名单核对"] not in {"是", "否"}:
            _issue(data.issues, DOCTOR_SHEET, row_number, "CA名单核对", "必须从模板下拉值中选择")
        if review == "通过" and not ca_matched:
            _issue(data.issues, DOCTOR_SHEET, row_number, "CA名单核对", "医务审核通过的开方医师必须完成 CA 名单核对")

        years = _optional_int(values["执业年限"], data.issues, DOCTOR_SHEET, row_number, "执业年限", 0, 60)
        fee = _optional_money(values["挂号费（元）"], data.issues, DOCTOR_SHEET, row_number, "挂号费（元）")
        row_errors = [x for x in data.errors if x.sheet == DOCTOR_SHEET and x.row == row_number]
        if not row_errors:
            data.doctors.append(DoctorImportRow(
                row_number=row_number,
                phone=phone,
                name=values["姓名"][:64],
                id_card=id_card,
                license_no=values["医师资格证号"][:64],
                practice_no=values["医师执业证号"][:64],
                dept=values["科室名称"][:32],
                dept_code=dept_code[:10],
                subject_code=subject_code[:10],
                subject_name=subject_name[:30],
                title=values["职称"][:32],
                years=years,
                good_at=values["擅长"][:255] or None,
                register_fee_yuan=fee,
                review_result=review,
                ca_matched=ca_matched,
            ))


def _parse_pharmacists(ws, data: ParsedPersonnel) -> None:
    seen_usernames: dict[str, int] = {}
    seen_ids: dict[str, int] = {}
    for row_number, values in _sheet_rows(ws, PHARMACIST_HEADERS, PHARMACIST_REQUIRED, data.issues) or []:
        username = values["登录用户名"]
        id_card = values["身份证号"].upper()
        mobile = values["手机号"]
        if username and not _USERNAME_RE.fullmatch(username):
            _issue(data.issues, PHARMACIST_SHEET, row_number, "登录用户名", "须以英文字母开头，仅含字母、数字、点、横线或下划线，长度 3–64")
        if id_card and not valid_id_card(id_card):
            _issue(data.issues, PHARMACIST_SHEET, row_number, "身份证号", "格式、出生日期或校验位不正确")
        if mobile and not _PHONE_RE.fullmatch(mobile):
            _issue(data.issues, PHARMACIST_SHEET, row_number, "手机号", "格式不正确")
        if username in seen_usernames:
            _issue(data.issues, PHARMACIST_SHEET, row_number, "登录用户名", f"与第{seen_usernames[username]}行重复")
        elif username:
            seen_usernames[username] = row_number
        if id_card in seen_ids:
            _issue(data.issues, PHARMACIST_SHEET, row_number, "身份证号", f"与第{seen_ids[id_card]}行重复")
        elif id_card:
            seen_ids[id_card] = row_number

        state = values["启用状态"]
        review = values["资质审核结论"]
        ca_value = values["CA名单核对"]
        if state not in {"启用", "停用"}:
            _issue(data.issues, PHARMACIST_SHEET, row_number, "启用状态", "必须从模板下拉值中选择")
        if review not in {"待审核", "通过", "不通过"}:
            _issue(data.issues, PHARMACIST_SHEET, row_number, "资质审核结论", "必须从模板下拉值中选择")
        if ca_value not in {"是", "否"}:
            _issue(data.issues, PHARMACIST_SHEET, row_number, "CA名单核对", "必须从模板下拉值中选择")
        if state == "启用" and (review != "通过" or ca_value != "是"):
            _issue(data.issues, PHARMACIST_SHEET, row_number, "启用状态", "启用的审方药师必须资质审核通过且完成 CA 名单核对")

        row_errors = [x for x in data.errors if x.sheet == PHARMACIST_SHEET and x.row == row_number]
        if not row_errors:
            data.pharmacists.append(PharmacistImportRow(
                row_number=row_number,
                username=username,
                name=values["姓名"][:64],
                id_card=id_card,
                mobile=mobile or None,
                active=state == "启用",
                review_result=review,
                ca_matched=ca_value == "是",
            ))


def parse_personnel_workbook(path: Path, dictionary_dir: Path) -> ParsedPersonnel:
    data = ParsedPersonnel()
    if not path.is_file():
        _issue(data.issues, "工作簿", None, None, "文件不存在或不是普通文件")
        return data
    if path.suffix.lower() != ".xlsx":
        _issue(data.issues, "工作簿", None, None, "只接受 .xlsx 文件")
        return data
    if path.stat().st_size > MAX_WORKBOOK_BYTES:
        _issue(data.issues, "工作簿", None, None, "文件超过 10MB 限制")
        return data
    try:
        with ZipFile(path) as archive:
            if len(archive.infolist()) > 2000:
                _issue(data.issues, "工作簿", None, None, "内部文件数量异常")
                return data
            if sum(item.file_size for item in archive.infolist()) > MAX_WORKBOOK_UNCOMPRESSED_BYTES:
                _issue(data.issues, "工作簿", None, None, "解压后内容超过 50MB 限制")
                return data
    except (BadZipFile, OSError):
        _issue(data.issues, "工作簿", None, None, "不是有效的 xlsx 压缩结构")
        return data
    try:
        # read_only=True 对 ws.cell() 随机访问会反复扫描 XML；文件已有 10MB 上限，
        # 使用普通内存加载并保持 data_only=False 以识别、拒绝公式。
        wb = load_workbook(path, read_only=False, data_only=False, keep_links=False)
    except Exception:  # noqa: BLE001
        _issue(data.issues, "工作簿", None, None, "无法读取，请确认文件未损坏且来自最新版模板")
        return data
    try:
        for title in (DOCTOR_SHEET, PHARMACIST_SHEET):
            if title not in wb.sheetnames:
                _issue(data.issues, "工作簿", None, title, "缺少必需工作表")
        if data.errors:
            return data
        try:
            dept_codes = _load_code_map(dictionary_dir / "科室类别字典3.2.csv")
            subject_codes = _load_code_map(dictionary_dir / "诊疗科目字典3.1.csv")
        except OSError:
            _issue(data.issues, "字典", None, None, "服务器监管字典文件缺失")
            return data
        _parse_doctors(wb[DOCTOR_SHEET], data, dept_codes, subject_codes)
        _parse_pharmacists(wb[PHARMACIST_SHEET], data)
        if NURSE_SHEET in wb.sheetnames:
            ws = wb[NURSE_SHEET]
            data.nurse_rows = sum(
                1 for row in ws.iter_rows(min_row=2, values_only=True)
                if any(value not in (None, "") for value in row)
            )
            if data.nurse_rows:
                _issue(
                    data.issues, NURSE_SHEET, None, None,
                    f"检测到 {data.nurse_rows} 行，仅计数留档，本次不会导入护士账号",
                    level="WARN",
                )
    finally:
        wb.close()

    cross_ids: dict[str, tuple[str, int]] = {}
    for sheet, rows in ((DOCTOR_SHEET, data.doctors), (PHARMACIST_SHEET, data.pharmacists)):
        for row in rows:
            previous = cross_ids.get(row.id_card)
            if previous:
                _issue(data.issues, sheet, row.row_number, "身份证号", f"与{previous[0]}第{previous[1]}行重复")
            else:
                cross_ids[row.id_card] = (sheet, row.row_number)
    return data


def _safe_decrypt(value: str | None) -> str | None:
    if not value:
        return None
    try:
        return decrypt(value)
    except Exception:  # noqa: BLE001
        return None


def build_database_plan(
    data: ParsedPersonnel,
    doctor_accounts: list[tuple[Any, Any]],
    existing_staff: list[Any],
) -> DatabasePlan:
    """用已加载的 ORM 记录构建计划；函数本身不写库，便于测试和 dry-run。"""
    plan = DatabasePlan()
    phone_map: dict[str, list[Any]] = {}
    existing_ids: dict[str, list[tuple[str, int]]] = {}
    bad_doctor_phones = bad_doctor_ids = bad_staff_ids = 0
    for doctor, user in doctor_accounts:
        phone = _safe_decrypt(getattr(user, "phone_enc", None))
        if getattr(user, "phone_enc", None) and (not phone or not _PHONE_RE.fullmatch(phone)):
            bad_doctor_phones += 1
        elif phone:
            phone_map.setdefault(phone, []).append(doctor)
        id_card = _safe_decrypt(getattr(doctor, "id_card_enc", None))
        if getattr(doctor, "id_card_enc", None) and (not id_card or not valid_id_card(id_card)):
            bad_doctor_ids += 1
        elif id_card:
            existing_ids.setdefault(id_card.upper(), []).append(("doctor", int(doctor.id)))

    staff_map = {str(item.username): item for item in existing_staff}
    for staff in existing_staff:
        id_card = _safe_decrypt(getattr(staff, "id_card_enc", None))
        if getattr(staff, "id_card_enc", None) and (not id_card or not valid_id_card(id_card)):
            bad_staff_ids += 1
        elif id_card:
            existing_ids.setdefault(id_card.upper(), []).append(("staff", int(staff.id)))
    if bad_doctor_phones:
        _issue(
            plan.issues, "数据库", None, "医生手机号",
            f"{bad_doctor_phones} 条已有密文无法用当前密钥解密或格式错误，禁止继续匹配",
        )
    if bad_doctor_ids:
        _issue(
            plan.issues, "数据库", None, "医生身份证",
            f"{bad_doctor_ids} 条已有密文无法用当前密钥解密或格式错误，禁止继续导入",
        )
    if bad_staff_ids:
        _issue(
            plan.issues, "数据库", None, "员工身份证",
            f"{bad_staff_ids} 条已有密文无法用当前密钥解密或格式错误，禁止继续导入",
        )

    for row in data.doctors:
        matches = phone_map.get(row.phone, [])
        if not matches:
            _issue(
                plan.issues, DOCTOR_SHEET, row.row_number, "手机号",
                "未找到首次登录医生档案；请本人先完成微信手机号授权登录",
            )
            continue
        if len(matches) > 1:
            _issue(plan.issues, DOCTOR_SHEET, row.row_number, "手机号", "数据库存在多个匹配账号，必须人工清理")
            continue
        target = matches[0]
        conflicts = [
            f"{kind}/{record_id}" for kind, record_id in existing_ids.get(row.id_card, [])
            if not (kind == "doctor" and record_id == int(target.id))
        ]
        if conflicts:
            _issue(
                plan.issues, DOCTOR_SHEET, row.row_number, "身份证号",
                f"已绑定其他数据库主体（内部标识：{','.join(conflicts)}）",
            )
            continue
        plan.doctors.append(DoctorOperation(row=row, target=target))

    for row in data.pharmacists:
        target = staff_map.get(row.username)
        if target is not None and getattr(target, "role", None) != "pharmacist":
            _issue(
                plan.issues, PHARMACIST_SHEET, row.row_number, "登录用户名",
                "已被非药师角色占用，禁止自动改权",
            )
            continue
        conflicts = [
            f"{kind}/{record_id}" for kind, record_id in existing_ids.get(row.id_card, [])
            if not (target is not None and kind == "staff" and record_id == int(target.id))
        ]
        if conflicts:
            _issue(
                plan.issues, PHARMACIST_SHEET, row.row_number, "身份证号",
                f"已绑定其他数据库主体（内部标识：{','.join(conflicts)}）",
            )
            continue
        plan.pharmacists.append(PharmacistOperation(row=row, target=target))
    return plan


async def apply_database_plan(
    db,
    plan: DatabasePlan,
    *,
    approve_doctors: bool,
    source_digest: str | None = None,
) -> ApplyResult:
    credentials: list[tuple[str, str]] = []
    doctors_approved = 0
    pharmacists_created = pharmacists_updated = 0
    for operation in plan.doctors:
        row, doctor = operation.row, operation.target
        doctor.name = row.name
        doctor.id_card_enc = encrypt(row.id_card)
        doctor.license_no = row.license_no
        doctor.practice_no = row.practice_no
        doctor.dept = row.dept
        doctor.dept_code = row.dept_code
        doctor.subject_code = row.subject_code
        doctor.subject_name = row.subject_name
        doctor.title = row.title
        if row.years is not None:
            doctor.years = row.years
        if row.good_at is not None:
            doctor.good_at = row.good_at
        if row.register_fee_yuan is not None:
            doctor.register_fee_fen = int(row.register_fee_yuan * 100)
        if approve_doctors:
            if row.review_result == "通过" and row.ca_matched:
                doctor.audit_status = "approved"
                doctors_approved += 1
            elif row.review_result == "不通过":
                doctor.audit_status = "rejected"
            else:
                doctor.audit_status = "pending"

    for operation in plan.pharmacists:
        row = operation.row
        if operation.is_new:
            password = generate_temporary_password()
            staff = Staff(
                username=row.username,
                password_hash=hash_password(password),
                role="pharmacist",
                name=row.name,
                active=row.active,
                id_card_enc=encrypt(row.id_card),
            )
            db.add(staff)
            credentials.append((row.username, password))
            pharmacists_created += 1
        else:
            staff = operation.target
            staff.name = row.name
            staff.active = row.active
            staff.id_card_enc = encrypt(row.id_card)
            pharmacists_updated += 1

    db.add(AuditLog(
        actor_id=0,
        actor_name="离线人员导入",
        actor_role="system",
        action="批量导入人员",
        target_type="personnel_import",
        detail=(
            f"医生更新{len(plan.doctors)}，医生终审{doctors_approved}，"
            f"药师新增{pharmacists_created}，药师更新{pharmacists_updated}，"
            f"工作簿SHA256={source_digest or '未记录'}"
        ),
    ))
    await db.flush()
    return ApplyResult(
        doctors_updated=len(plan.doctors),
        doctors_approved=doctors_approved,
        pharmacists_created=pharmacists_created,
        pharmacists_updated=pharmacists_updated,
        credentials=tuple(credentials),
    )


def write_credentials_file(path: Path, credentials: tuple[tuple[str, str], ...]) -> None:
    """以排他方式创建 0600 凭据文件；绝不覆盖已有文件。"""
    if not credentials:
        return
    if not path.is_absolute():
        raise ValueError("凭据输出路径必须是绝对路径")
    if not path.parent.is_dir():
        raise ValueError("凭据输出目录不存在")
    if os.name == "posix" and path.parent.stat().st_mode & 0o077:
        raise PermissionError("凭据输出目录权限过宽，必须先 chmod 700")
    flags = os.O_WRONLY | os.O_CREAT | os.O_EXCL
    fd = os.open(path, flags, 0o600)
    try:
        with os.fdopen(fd, "w", encoding="utf-8-sig", newline="") as fp:
            writer = csv.writer(fp)
            writer.writerow(["登录用户名", "随机初始密码"])
            writer.writerows(credentials)
        os.chmod(path, 0o600)
    except Exception:
        try:
            path.unlink(missing_ok=True)
        finally:
            raise
