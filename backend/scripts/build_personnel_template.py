"""生成医生、药师和护士人员资料收集模板（模板本身不含个人信息）。

用法：
    python -m scripts.build_personnel_template

填写后的文件包含身份证号等敏感信息，不得放入仓库或通过聊天传输。
"""
from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName


BACKEND_DIR = Path(__file__).resolve().parents[1]
REPO_DIR = BACKEND_DIR.parent
DICT_DIR = BACKEND_DIR / "data" / "tj_dicts"
DEFAULT_OUTPUT = REPO_DIR / "docs" / "templates" / "人员批量导入模板.xlsx"

HEADER_FILL_REQUIRED = PatternFill("solid", fgColor="C62828")
HEADER_FILL_OPTIONAL = PatternFill("solid", fgColor="455A64")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FILL = PatternFill("solid", fgColor="DCE6F1")
WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")


DOCTOR_COLUMNS = [
    ("手机号", True, "本人首次登录医生端时授权的手机号；用于匹配微信账号，11 位数字。"),
    ("姓名", True, "必须与身份证、医师证书及放心签个人证书主体逐字一致。"),
    ("身份证号", True, "18 位；导入后加密保存，不回显。"),
    ("医师资格证号", True, "按医师资格证原件填写，按文本保存。"),
    ("医师执业证号", True, "按医师执业证原件填写，按文本保存。"),
    ("科室名称", True, "系统展示名称，如内科、妇产科、中医科。"),
    ("科室类别编码", True, "天津监管字典 3.2 的 deptID。"),
    ("诊疗科目编码", True, "天津监管字典 3.1 的 subjectCode，须在医院许可范围内。"),
    ("诊疗科目名称", True, "必须与诊疗科目编码对应的字典名称一致。"),
    ("职称", True, "如主任医师、副主任医师、主治医师。"),
    ("执业年限", False, "整数；无法确认可暂留空。"),
    ("擅长", False, "患者端展示内容，不含广告或绝对化表述。"),
    ("挂号费（元）", False, "非负金额；留空时按系统默认值处理。"),
    ("医务审核结论", True, "只有“通过”的人员才允许终审通过。"),
    ("CA名单核对", True, "姓名及身份证已与最新版个人数字证书回函核对。"),
    ("备注", False, "不得粘贴证件照片、账号密码或其他无关敏感信息。"),
]

PHARMACIST_COLUMNS = [
    ("登录用户名", True, "运营后台登录名；建议使用单位内部工号，不使用身份证号。"),
    ("姓名", True, "必须与身份证、药师证书及放心签个人证书主体逐字一致。"),
    ("身份证号", True, "18 位；导入后加密保存，不回显。"),
    ("手机号", False, "仅用于内部联络，当前版本不写入 staff 表。"),
    ("药学资格证号", False, "医院资质留档字段，当前版本暂不写入 staff 表。"),
    ("执业药师注册证号", False, "如适用；当前版本暂不写入 staff 表。"),
    ("科室/部门", False, "如药剂科、审方中心。"),
    ("职称", False, "如主管药师、药师。"),
    ("启用状态", True, "选择“启用”或“停用”。"),
    ("资质审核结论", True, "只有“通过”的人员才创建为可用审方账号。"),
    ("CA名单核对", True, "姓名及身份证已与最新版个人数字证书回函核对。"),
    ("备注", False, "初始密码不填在本表；批量导入时单独安全生成和分发。"),
]

NURSE_COLUMNS = [
    ("姓名", True, "按身份证及护士执业证书填写。"),
    ("身份证号", True, "18 位；本表当前只收集，不导入生产系统。"),
    ("手机号", True, "内部联络手机号。"),
    ("科室/部门", True, "实际执业科室。"),
    ("职称", False, "如主管护师、护师、护士。"),
    ("护士执业证书编号", True, "按证书原件填写，按文本保存。"),
    ("专业技术资格证号", False, "如有则填写。"),
    ("执业范围", False, "按备案内容填写。"),
    ("证书有效期起", False, "格式 YYYY-MM-DD。"),
    ("证书有效期止", False, "格式 YYYY-MM-DD。"),
    ("在岗状态", True, "选择“在岗”或“停用”。"),
    ("备注", False, "护士当前无系统角色，不得借用药师或管理员权限。"),
]


def _read_csv(name: str) -> list[list[str]]:
    with (DICT_DIR / name).open("r", encoding="utf-8-sig", newline="") as fp:
        return [row for row in csv.reader(fp) if row]


def _setup_data_sheet(ws, columns: list[tuple[str, bool, str]]) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}501"
    ws.row_dimensions[1].height = 36
    for col_idx, (name, required, help_text) in enumerate(columns, 1):
        cell = ws.cell(1, col_idx, f"{name}{' *' if required else ''}")
        cell.fill = HEADER_FILL_REQUIRED if required else HEADER_FILL_OPTIONAL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.comment = Comment(help_text, "逑贝")
        width = max(14, min(30, len(name) * 2 + 8))
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        for row_idx in range(2, 502):
            ws.cell(row_idx, col_idx).number_format = "@"
            ws.cell(row_idx, col_idx).alignment = Alignment(vertical="top", wrap_text=True)


def _add_list_validation(ws, header_name: str, values: list[str]) -> None:
    headers = {str(cell.value).replace(" *", ""): cell.column for cell in ws[1]}
    col = headers[header_name]
    dv = DataValidation(type="list", formula1=f'"{",".join(values)}"', allow_blank=False)
    dv.error = "请选择下拉列表中的有效值"
    dv.errorTitle = "值不合法"
    ws.add_data_validation(dv)
    dv.add(f"{get_column_letter(col)}2:{get_column_letter(col)}501")


def _add_named_validation(wb, ws, header_name: str, range_name: str) -> None:
    headers = {str(cell.value).replace(" *", ""): cell.column for cell in ws[1]}
    col = headers[header_name]
    dv = DataValidation(type="list", formula1=f"={range_name}", allow_blank=False)
    dv.error = "请选择字典工作表中的有效编码"
    dv.errorTitle = "编码不合法"
    ws.add_data_validation(dv)
    dv.add(f"{get_column_letter(col)}2:{get_column_letter(col)}501")


def _add_dictionary_sheet(wb, title: str, rows: list[list[str]]) -> None:
    ws = wb.create_sheet(title)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = HEADER_FILL_OPTIONAL
        cell.font = HEADER_FONT
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 36
    for row in ws.iter_rows():
        for cell in row:
            cell.number_format = "@"


def build_template(output: Path = DEFAULT_OUTPUT) -> Path:
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    intro = wb.active
    intro.title = "填写说明"
    intro.column_dimensions["A"].width = 22
    intro.column_dimensions["B"].width = 105
    intro.append(["项目", "说明"])
    intro.append(["模板版本", "1.0（2026-07-25）"])
    intro.append(["红色表头", "必填；灰色表头为选填。每人一行，不合并单元格。"])
    intro.append(["医生导入前置", "医生本人先用微信手机号登录医生端一次，系统生成待审核账号；导入时按手机号匹配，禁止预造 openid。"])
    intro.append(["药师账号", "表内不保存初始密码；导入时生成独立强密码，通过受控渠道逐人分发并要求首次使用后重置。"])
    intro.append(["护士范围", "当前系统没有护士角色，天津监管护理类接口也不在本院对接范围；护士页只作花名册，暂不导入。"])
    intro.append(["CA一致性", "参与开方或审方的姓名、身份证必须与放心签最新版个人数字证书回函逐字一致。"])
    intro.append(["编码", "医生的诊疗科目和科室类别必须从本模板字典页选择，并在医院互联网诊疗许可范围内。"])
    intro.append(["敏感信息", "填写后的文件不得提交 Git、不得发到聊天或普通邮件；应加密传输并放入服务器 chmod 600 的受控目录。"])
    intro.append(["导入结果", "身份证仅加密落库，密码仅保存哈希；导入后运行监管预检，再由本人完成放心签双录。"])
    for cell in intro[1]:
        cell.fill = HEADER_FILL_OPTIONAL
        cell.font = HEADER_FONT
    for row in intro.iter_rows(min_row=2):
        row[0].fill = SECTION_FILL
        row[0].font = Font(bold=True)
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
    intro["A9"].fill = WARNING_FILL
    intro["B9"].fill = WARNING_FILL
    intro.freeze_panes = "A2"

    doctors = wb.create_sheet("医生")
    pharmacists = wb.create_sheet("药师")
    nurses = wb.create_sheet("护士_暂不导入")
    _setup_data_sheet(doctors, DOCTOR_COLUMNS)
    _setup_data_sheet(pharmacists, PHARMACIST_COLUMNS)
    _setup_data_sheet(nurses, NURSE_COLUMNS)

    _add_list_validation(doctors, "医务审核结论", ["待审核", "通过", "不通过"])
    _add_list_validation(doctors, "CA名单核对", ["是", "否"])
    _add_list_validation(pharmacists, "启用状态", ["启用", "停用"])
    _add_list_validation(pharmacists, "资质审核结论", ["待审核", "通过", "不通过"])
    _add_list_validation(pharmacists, "CA名单核对", ["是", "否"])
    _add_list_validation(nurses, "在岗状态", ["在岗", "停用"])

    dept_rows = _read_csv("科室类别字典3.2.csv")
    subject_rows = _read_csv("诊疗科目字典3.1.csv")
    _add_dictionary_sheet(wb, "科室类别字典3.2", dept_rows)
    _add_dictionary_sheet(wb, "诊疗科目字典3.1", subject_rows)
    wb.defined_names.add(
        DefinedName("dept_codes", attr_text=f"'科室类别字典3.2'!$A$2:$A${len(dept_rows)}")
    )
    wb.defined_names.add(
        DefinedName("subject_codes", attr_text=f"'诊疗科目字典3.1'!$A$2:$A${len(subject_rows)}")
    )
    _add_named_validation(wb, doctors, "科室类别编码", "dept_codes")
    _add_named_validation(wb, doctors, "诊疗科目编码", "subject_codes")

    wb.save(output)
    # 立即回读，避免交付损坏的二进制模板。
    check = load_workbook(output, read_only=True, data_only=False)
    expected = {"填写说明", "医生", "药师", "护士_暂不导入", "科室类别字典3.2", "诊疗科目字典3.1"}
    if set(check.sheetnames) != expected:
        raise RuntimeError("生成的人员模板工作表不完整")
    check.close()
    return output


if __name__ == "__main__":
    print(build_template())
